#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para gerar arquivo Excel de teste para importação de alunos
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Alunos"

# Definir cabeçalhos
headers = [
    "Nome",
    "Data de Nascimento",
    "RG",
    "CPF",
    "Endereço",
    "Número",
    "Bairro",
    "Município",
    "Estado",
    "Escola",
    "Tipo Escola",
    "Série",
    "Turno",
    "Número de Pessoas na Casa",
    "Contato 1",
    "Contato 2"
]

# Estilizar cabeçalho
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Escrever cabeçalhos
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Dados de exemplo
nomes = [
    "João Silva Santos",
    "Maria Oliveira Costa",
    "Pedro Almeida Souza",
    "Ana Paula Ferreira",
    "Carlos Eduardo Lima",
    "Juliana Rodrigues",
    "Fernando Henrique Martins",
    "Beatriz Araújo",
    "Lucas Gabriel Pereira",
    "Isabela Cristina Nunes"
]

ruas = [
    "Rua das Flores",
    "Avenida Principal",
    "Rua do Comércio",
    "Avenida Central",
    "Rua da Paz",
    "Avenida dos Trabalhadores",
    "Rua São Paulo",
    "Avenida Brasil",
    "Rua das Palmeiras",
    "Avenida Independência"
]

bairros = [
    "Centro",
    "Jardim Primavera",
    "Vila Nova",
    "Bela Vista",
    "São José",
    "Parque Industrial",
    "Alto da Boa Vista",
    "Nova Esperança",
    "Santa Maria",
    "Vila Esperança"
]

municipios = [
    "São Paulo",
    "Rio de Janeiro",
    "Belo Horizonte",
    "Curitiba",
    "Porto Alegre",
    "Brasília",
    "Salvador",
    "Fortaleza",
    "Recife",
    "Goiânia"
]

estados = ["SP", "RJ", "MG", "PR", "RS", "DF", "BA", "CE", "PE", "GO"]

escolas = [
    "Escola Municipal João Silva",
    "Colégio Estadual Maria Santos",
    "Escola Particular Dom Pedro",
    "Colégio Público Ana Paula",
    "Instituto Educacional Carlos",
    "Escola Municipal Juliana",
    "Colégio Estadual Fernando",
    "Escola Particular Beatriz",
    "Colégio Público Lucas",
    "Instituto Educacional Isabela"
]

tipos_escola = ["Pública", "Privada"]
series = ["1º Ano", "2º Ano", "3º Ano", "4º Ano", "5º Ano", "6º Ano", "7º Ano", "8º Ano", "9º Ano"]
turnos = ["Matutino", "Vespertino", "Noturno", "Integral"]

# Gerar dados de exemplo
for row in range(2, 12):  # 10 alunos (linhas 2 a 11)
    nome_idx = row - 2
    nome = nomes[nome_idx]
    
    # Data de nascimento (entre 6 e 18 anos atrás)
    anos_atras = random.randint(6, 18)
    data_nasc = datetime.now() - timedelta(days=anos_atras * 365 + random.randint(0, 365))
    # Salvar como objeto datetime para que o Excel reconheça como data
    
    # RG (formato: XX.XXX.XXX-X)
    rg = f"{random.randint(10, 99)}.{random.randint(100, 999)}.{random.randint(100, 999)}-{random.randint(1, 9)}"
    
    # CPF (formato: XXX.XXX.XXX-XX)
    cpf = f"{random.randint(100, 999)}.{random.randint(100, 999)}.{random.randint(100, 999)}-{random.randint(10, 99)}"
    
    # Endereço
    endereco = ruas[random.randint(0, len(ruas) - 1)]
    numero = str(random.randint(1, 9999))
    bairro = bairros[random.randint(0, len(bairros) - 1)]
    municipio = municipios[random.randint(0, len(municipios) - 1)]
    estado = estados[random.randint(0, len(estados) - 1)]
    
    # Escola
    escola = escolas[nome_idx]
    tipo_escola = tipos_escola[random.randint(0, 1)]
    serie = series[random.randint(0, len(series) - 1)]
    turno = turnos[random.randint(0, len(turnos) - 1)]
    
    # Número de pessoas na casa
    num_pessoas = random.randint(2, 8)
    
    # Contatos
    contato1 = f"({random.randint(10, 99)}) {random.randint(90000, 99999)}-{random.randint(1000, 9999)}"
    contato2 = f"({random.randint(10, 99)}) {random.randint(90000, 99999)}-{random.randint(1000, 9999)}" if random.random() > 0.3 else ""
    
    # Escrever dados
    # Escrever nome
    ws.cell(row=row, column=1, value=nome).alignment = Alignment(horizontal="left", vertical="center")
    
    # Escrever data de nascimento como objeto datetime do Excel
    cell_data = ws.cell(row=row, column=2, value=data_nasc)
    cell_data.alignment = Alignment(horizontal="left", vertical="center")
    # Formatar a célula para mostrar a data no formato dd/mm/yyyy
    cell_data.number_format = 'dd/mm/yyyy'
    
    # Escrever os demais campos
    data = [
        rg,
        cpf,
        endereco,
        numero,
        bairro,
        municipio,
        estado,
        escola,
        tipo_escola,
        serie,
        turno,
        num_pessoas,
        contato1,
        contato2
    ]
    
    for col_idx, value in enumerate(data, start=3):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Ajustar largura das colunas
column_widths = {
    'A': 25,  # Nome
    'B': 18,  # Data de Nascimento
    'C': 15,  # RG
    'D': 18,  # CPF
    'E': 25,  # Endereço
    'F': 10,  # Número
    'G': 20,  # Bairro
    'H': 20,  # Município
    'I': 10,  # Estado
    'J': 30,  # Escola
    'K': 15,  # Tipo Escola
    'L': 12,  # Série
    'M': 15,  # Turno
    'N': 25,  # Número de Pessoas na Casa
    'O': 18,  # Contato 1
    'P': 18   # Contato 2
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Congelar primeira linha
ws.freeze_panes = 'A2'

# Salvar arquivo
filename = "exemplo_importacao_alunos.xlsx"
wb.save(filename)
print(f"Arquivo Excel criado com sucesso: {filename}")
print(f"Total de alunos: {len(nomes)}")
print(f"Colunas: {len(headers)}")
print("\nIMPORTANTE: Este arquivo contem dados de exemplo validos.")
print("Todos os registros devem ser importados com sucesso (importacao atomica).")

