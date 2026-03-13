#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para gerar arquivo Excel COM ERROS para testar validação atômica
Este arquivo deve FALHAR na importação (nenhum aluno será importado)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from datetime import datetime, timedelta
import random

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Alunos"

# Definir cabeçalhos
headers = [
    "Nome",
    "Data de Nascimento",
    "RG",
    "CPF",
    "Endereço",
    "Número",
    "Bairro",
    "Município",
    "Estado",
    "Escola",
    "Tipo Escola",
    "Série",
    "Turno",
    "Número de Pessoas na Casa",
    "Contato 1",
    "Contato 2"
]

# Estilizar cabeçalho
header_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Escrever cabeçalhos
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Dados de exemplo - alguns válidos, alguns com erros
dados = [
    # Linha 2 - VÁLIDA
    ["João Silva Santos", "15/03/2010", "12.345.678-9", "123.456.789-00", 
     "Rua das Flores", "123", "Centro", "São Paulo", "SP", 
     "Escola Municipal João Silva", "Pública", "5º Ano", "Matutino", "4", 
     "(11) 98765-4321", "(11) 91234-5678"],
    
    # Linha 3 - ERRO: Nome vazio
    ["", "20/05/2011", "23.456.789-0", "234.567.890-11", 
     "Avenida Principal", "456", "Jardim Primavera", "Rio de Janeiro", "RJ", 
     "Colégio Estadual Maria", "Privada", "4º Ano", "Vespertino", "3", 
     "(21) 98765-4321", ""],
    
    # Linha 4 - VÁLIDA
    ["Pedro Almeida Souza", "10/08/2009", "34.567.890-1", "345.678.901-22", 
     "Rua do Comércio", "789", "Vila Nova", "Belo Horizonte", "MG", 
     "Escola Particular Dom Pedro", "Privada", "6º Ano", "Integral", "5", 
     "(31) 98765-4321", "(31) 91234-5678"],
    
    # Linha 5 - ERRO: Data inválida
    ["Ana Paula Ferreira", "32/13/2012", "45.678.901-2", "456.789.012-33", 
     "Avenida Central", "321", "Bela Vista", "Curitiba", "PR", 
     "Colégio Público Ana", "Pública", "3º Ano", "Noturno", "2", 
     "(41) 98765-4321", ""],
    
    # Linha 6 - ERRO: CPF vazio
    ["Carlos Eduardo Lima", "25/12/2010", "56.789.012-3", "", 
     "Rua da Paz", "654", "São José", "Porto Alegre", "RS", 
     "Instituto Educacional Carlos", "Privada", "5º Ano", "Matutino", "6", 
     "(51) 98765-4321", "(51) 91234-5678"],
    
    # Linha 7 - ERRO: Estado com mais de 2 caracteres
    ["Juliana Rodrigues", "05/07/2011", "67.890.123-4", "567.890.123-44", 
     "Avenida dos Trabalhadores", "987", "Parque Industrial", "Brasília", "DFA", 
     "Escola Municipal Juliana", "Pública", "4º Ano", "Vespertino", "4", 
     "(61) 98765-4321", ""],
    
    # Linha 8 - ERRO: Tipo de escola inválido
    ["Fernando Henrique Martins", "18/09/2009", "78.901.234-5", "678.901.234-55", 
     "Rua São Paulo", "147", "Alto da Boa Vista", "Salvador", "BA", 
     "Colégio Estadual Fernando", "Particular", "6º Ano", "Integral", "7", 
     "(71) 98765-4321", "(71) 91234-5678"],
    
    # Linha 9 - ERRO: Turno inválido
    ["Beatriz Araújo", "22/11/2010", "89.012.345-6", "789.012.345-66", 
     "Avenida Brasil", "258", "Nova Esperança", "Fortaleza", "CE", 
     "Escola Particular Beatriz", "Privada", "5º Ano", "Diurno", "3", 
     "(85) 98765-4321", ""],
    
    # Linha 10 - ERRO: Número de pessoas inválido (0)
    ["Lucas Gabriel Pereira", "30/01/2011", "90.123.456-7", "890.123.456-77", 
     "Rua das Palmeiras", "369", "Santa Maria", "Recife", "PE", 
     "Colégio Público Lucas", "Pública", "4º Ano", "Matutino", "0", 
     "(81) 98765-4321", "(81) 91234-5678"],
]

# Escrever dados
for row, data in enumerate(dados, start=2):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Destacar linhas com erro (linhas ímpares após a primeira)
        if row in [3, 5, 6, 7, 8, 9, 10]:
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

# Ajustar largura das colunas
column_widths = {
    'A': 25,  'B': 18,  'C': 15,  'D': 18,  'E': 25,  'F': 10,
    'G': 20,  'H': 20,  'I': 10,  'J': 30,  'K': 15,  'L': 12,
    'M': 15,  'N': 25,  'O': 18,  'P': 18
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Congelar primeira linha
ws.freeze_panes = 'A2'

# Adicionar comentário na primeira célula
ws['A1'].comment = Comment("Este arquivo contem erros propositais para testar validacao atomica", "Sistema")

# Salvar arquivo
filename = "exemplo_importacao_com_erros.xlsx"
wb.save(filename)
print(f"Arquivo Excel COM ERROS criado: {filename}")
print(f"Total de linhas: {len(dados)}")
print(f"Linhas com erro: 3, 5, 6, 7, 8, 9, 10")
print("\nIMPORTANTE: Este arquivo contem erros propositais.")
print("A importacao deve FALHAR e NENHUM aluno deve ser importado (validacao atomica).")
print("\nErros propositais:")
print("- Linha 3: Nome vazio")
print("- Linha 5: Data invalida (32/13/2012)")
print("- Linha 6: CPF vazio")
print("- Linha 7: Estado com mais de 2 caracteres (DFA)")
print("- Linha 8: Tipo de escola invalido (Particular)")
print("- Linha 9: Turno invalido (Diurno)")
print("- Linha 10: Numero de pessoas invalido (0)")

